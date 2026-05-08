"""
evaluation/excel_rankings.py

Excel ranking workbook writers for post-validation outputs.

Flag semantics:
- Red:    metric <= mean - std
- Yellow: within user-defined worst percentile
- Orange: both red and yellow
"""

from __future__ import annotations

import math
import statistics
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from evaluation.record_utils import truncate_text


HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
RED_FILL = PatternFill("solid", fgColor="F4CCCC")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
ORANGE_FILL = PatternFill("solid", fgColor="FCE5CD")

DEFAULT_YELLOW_RATIO = 0.10


def safe_float(value: Any, default: float = 0.0) -> float:
    """
    Convert a value to float safely.

    Args:
        value: Numeric-like value.
        default: Fallback value.

    Returns:
        Converted float, or default when conversion fails.
    """
    if value is None:
        return default

    try:
        number = float(value)
    except (TypeError, ValueError):
        return default

    if math.isnan(number) or math.isinf(number):
        return default

    return number


def clamp_ratio(ratio: float) -> float:
    """
    Clamp a ratio into [0.0, 1.0].

    Args:
        ratio: User-defined percentile ratio.

    Returns:
        Clamped ratio.
    """
    return max(0.0, min(1.0, safe_float(ratio, DEFAULT_YELLOW_RATIO)))


def percentile_count(total_count: int, ratio: float) -> int:
    """
    Convert a ratio into a selected row count.

    Args:
        total_count: Number of records.
        ratio: Worst-case ratio.

    Returns:
        Number of rows selected by the ratio.
    """
    ratio = clamp_ratio(ratio)

    if total_count <= 0 or ratio <= 0.0:
        return 0

    return max(1, math.ceil(total_count * ratio))


def add_rank(records: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Add 1-based rank values to already-sorted records.

    Args:
        records: Sorted records.

    Returns:
        Records with rank fields.
    """
    return [
        {"rank": rank, **record}
        for rank, record in enumerate(records, start=1)
    ]


def write_sheet(
    workbook: Workbook,
    sheet_name: str,
    records: Sequence[Dict[str, Any]],
    columns: Sequence[str],
) -> Worksheet:
    """
    Write one worksheet from records.

    Args:
        workbook: Target workbook.
        sheet_name: Worksheet name.
        records: Row records.
        columns: Ordered column names.

    Returns:
        Created worksheet.
    """
    sheet = workbook.create_sheet(title=sheet_name)
    sheet.append(list(columns))

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for record in records:
        sheet.append([record.get(column) for column in columns])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_index, column_name in enumerate(columns, start=1):
        column_letter = get_column_letter(column_index)
        width = 14

        if column_name in {"prediction", "reference", "pred_preview", "ref_preview"}:
            width = 48
        elif column_name in {"record_index", "dialog_index", "turn_index"}:
            width = 14
        elif column_name in {
            "confidence",
            "margin",
            "rouge_l",
            "bleu_2",
            "bertscore_f1",
            "red_threshold",
        }:
            width = 16
        elif column_name == "flag":
            width = 28

        sheet.column_dimensions[column_letter].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    return sheet


def mark_flag_cells(sheet: Worksheet, flag_column_name: str = "flag") -> None:
    """
    Apply color fill to the flag column.

    Args:
        sheet: Worksheet to format.
        flag_column_name: Header name of the flag column.
    """
    headers = [cell.value for cell in sheet[1]]

    if flag_column_name not in headers:
        return

    flag_col = headers.index(flag_column_name) + 1

    for row_index in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row_index, column=flag_col)
        value = str(cell.value or "").lower()

        if "orange" in value:
            cell.fill = ORANGE_FILL
        elif "red" in value:
            cell.fill = RED_FILL
        elif "yellow" in value:
            cell.fill = YELLOW_FILL


def format_numeric_columns(sheet: Worksheet, columns: Iterable[str]) -> None:
    """
    Apply consistent number formatting.

    Args:
        sheet: Worksheet to format.
        columns: Column names to format.
    """
    headers = [cell.value for cell in sheet[1]]

    for column_name in columns:
        if column_name not in headers:
            continue

        column_index = headers.index(column_name) + 1

        for row_index in range(2, sheet.max_row + 1):
            sheet.cell(row=row_index, column=column_index).number_format = "0.000000"


def metric_value(record: Dict[str, Any], metric_name: str) -> float:
    """
    Read a generation metric from a nested generation record.

    Args:
        record: Generation prediction record.
        metric_name: Metric key.

    Returns:
        Metric value as float. Missing values are treated as 0.
    """
    metrics = record.get("metrics", {})

    if not isinstance(metrics, dict):
        return 0.0

    return safe_float(metrics.get(metric_name), default=0.0)


def mean_std_arith(values: Sequence[float], subtraction=True) -> float:
    """
    Compute mean - population standard deviation.

    Args:
        values: Numeric values.

    Returns:
        Mean - std threshold.
    """
    if not values:
        return 0.0

    mean_value = statistics.mean(values)
    std_value = statistics.pstdev(values) if len(values) > 1 else 0.0

    if subtraction:
        return float(mean_value - std_value)
    else:
        return float(mean_value + std_value)


def build_flag(is_red: bool, is_yellow: bool) -> str:
    """
    Build one flag string from red/yellow conditions.

    Args:
        is_red: Whether the record satisfies the red condition.
        is_yellow: Whether the record satisfies the yellow condition.

    Returns:
        Flag string.
    """
    if is_red and is_yellow:
        return "orange_red_and_yellow"

    if is_red:
        return "red_mean_std_arith"

    if is_yellow:
        return "yellow_worst_percentile"

    return ""


def add_metric_flags(
    records: List[Dict[str, Any]],
    metric_name: str,
    red_threshold: float,
    yellow_ratio: float = DEFAULT_YELLOW_RATIO,
    lower_is_worse: bool = True,
) -> List[Dict[str, Any]]:
    """
    Add red/yellow/orange flags to sorted flat records.

    Args:
        records: Sorted flat records.
        metric_name: Metric column name.
        red_threshold: Red threshold, usually mean - std.
        yellow_ratio: User-defined worst percentile ratio.
        lower_is_worse: Whether lower metric values are worse.

    Returns:
        Records with flag fields.
    """
    yellow_count = percentile_count(len(records), yellow_ratio)
    output = []

    for index, record in enumerate(records):
        value = safe_float(record.get(metric_name), default=0.0)

        if lower_is_worse:
            is_red = value <= red_threshold
            is_yellow = index < yellow_count
        else:
            is_red = value >= red_threshold
            is_yellow = index < yellow_count

        output.append(
            {
                **record,
                "flag": build_flag(is_red=is_red, is_yellow=is_yellow),
                "red_threshold": red_threshold,
            }
        )

    return output

def add_bad_consensus_scores(
    records: List[Dict[str, Any]],
    thresholds: Dict[str, float],
) -> List[Dict[str, Any]]:
    """
    Add bad consensus score based on red-flag agreement.

    Red conditions:
    - rouge_l <= mean - std
    - bleu_2 <= mean - std
    - bertscore_f1 <= mean - std
    - nll >= mean + std
    - ppl >= mean + std

    Args:
        records: Flat generation records.
        thresholds: Metric red thresholds.

    Returns:
        Records with bad_consensus_score and bad_consensus_detail.
    """
    output = []

    for record in records:
        bad_votes = {
            "rouge_l_red": safe_float(record.get("rouge_l")) <= thresholds["rouge_l"],
            "bleu_2_red": safe_float(record.get("bleu_2")) <= thresholds["bleu_2"],
            "bertscore_f1_red": (
                safe_float(record.get("bertscore_f1")) <= thresholds["bertscore_f1"]
            ),
            "nll_red": safe_float(record.get("nll")) >= thresholds["nll"],
            "ppl_red": safe_float(record.get("ppl")) >= thresholds["ppl"],
        }

        bad_consensus_score = sum(1 for is_bad in bad_votes.values() if is_bad)
        bad_consensus_detail = ",".join(
            metric_name
            for metric_name, is_bad in bad_votes.items()
            if is_bad
        )

        output.append(
            {
                **record,
                "bad_consensus_score": bad_consensus_score,
                "bad_consensus_detail": bad_consensus_detail,
            }
        )

    return output


def flatten_classification_record(record: Dict[str, Any]) -> Dict[str, Any]:
    """
    Convert one classification record into flat Excel columns.

    Args:
        record: Unified classification prediction record.

    Returns:
        Flat row record.
    """
    sample_id = record.get("sample_id", {})

    if not isinstance(sample_id, dict):
        sample_id = {}

    return {
        "record_index": record.get("record_index"),
        "dialog_index": record.get("dialog_index", sample_id.get("dialog_index")),
        "turn_index": record.get("turn_index", sample_id.get("turn_index")),
        "label": record.get("label"),
        "pred": record.get("pred"),
        "confidence": safe_float(record.get("confidence"), default=0.0),
        "margin": safe_float(record.get("margin"), default=0.0),
        "is_correct": record.get("is_correct"),
    }


def save_classification_rankings_excel(
    records: List[Dict[str, Any]],
    output_path: Path,
    yellow_ratio: float = DEFAULT_YELLOW_RATIO,
) -> None:
    """
    Save classification ranking sheets.

    Flag semantics:
    - Wrong high confidence:
        red/yellow both point to dangerous wrong predictions.
    - Wrong low margin:
        red/yellow both point to uncertain wrong predictions.
    - Correct low margin:
        red/yellow both point to fragile correct predictions.
    - All by confidence:
        no red/yellow statistical flag, because correctness boundary is categorical.

    Args:
        records: Unified prediction records.
        output_path: Output xlsx path.
        yellow_ratio: User-defined worst percentile ratio.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)

    columns = [
        "rank",
        "record_index",
        "dialog_index",
        "turn_index",
        "label",
        "pred",
        "confidence",
        "margin",
        "is_correct",
        "flag",
        "red_threshold",
    ]

    flat_records = [flatten_classification_record(record) for record in records]

    wrong_records = [
        record for record in flat_records
        if record.get("is_correct") is False
    ]

    correct_records = [
        record for record in flat_records
        if record.get("is_correct") is True
    ]

    wrong_conf_threshold = mean_std_arith(
        [safe_float(record.get("confidence")) for record in wrong_records]
    )

    wrong_margin_threshold = mean_std_arith(
        [safe_float(record.get("margin")) for record in wrong_records]
    )

    correct_margin_threshold = mean_std_arith(
        [safe_float(record.get("margin")) for record in correct_records]
    )

    wrong_high_conf = sorted(
        wrong_records,
        key=lambda r: safe_float(r.get("confidence")),
        reverse=True,
    )

    wrong_low_margin = sorted(
        wrong_records,
        key=lambda r: safe_float(r.get("margin")),
        reverse=False,
    )

    correct_low_margin = sorted(
        correct_records,
        key=lambda r: safe_float(r.get("margin")),
        reverse=False,
    )

    all_by_confidence = sorted(
        flat_records,
        key=lambda r: safe_float(r.get("confidence")),
        reverse=True,
    )

    sheets = {
        "wrong_high_conf_desc": add_metric_flags(
            records=wrong_high_conf,
            metric_name="confidence",
            red_threshold=wrong_conf_threshold,
            yellow_ratio=yellow_ratio,
            lower_is_worse=False,
        ),
        "wrong_low_margin_asc": add_metric_flags(
            records=wrong_low_margin,
            metric_name="margin",
            red_threshold=wrong_margin_threshold,
            yellow_ratio=yellow_ratio,
            lower_is_worse=True,
        ),
        "correct_low_margin_asc": add_metric_flags(
            records=correct_low_margin,
            metric_name="margin",
            red_threshold=correct_margin_threshold,
            yellow_ratio=yellow_ratio,
            lower_is_worse=True,
        ),
        "all_by_confidence_desc": [
            {**record, "flag": "", "red_threshold": None}
            for record in all_by_confidence
        ],
    }

    for sheet_name, sheet_records in sheets.items():
        sheet = write_sheet(workbook, sheet_name, add_rank(sheet_records), columns)
        mark_flag_cells(sheet)
        format_numeric_columns(sheet, ["confidence", "margin", "red_threshold"])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def flatten_generation_record(record: Dict[str, Any]) -> Dict[str, Any]:
    """
    Convert nested generation metrics into Excel columns.

    Args:
        record: Unified generation prediction record.

    Returns:
        Flat row record for Excel.
    """
    rouge_l = metric_value(record, "rouge_l")
    bleu_2 = metric_value(record, "bleu_2")
    bertscore_f1 = metric_value(record, "bertscore_f1")
    ppl = metric_value(record, "ppl")
    nll = metric_value(record, "nll")

    return {
        "record_index": record.get("record_index"),
        "dialog_index": record.get("dialog_index"),
        "turn_index": record.get("turn_index"),
        "rouge_l": rouge_l,
        "bleu_2": bleu_2,
        "bertscore_f1": bertscore_f1,
        "nll": nll,
        "ppl": ppl,
        "pred_preview": truncate_text(record.get("prediction"), 180),
        "ref_preview": truncate_text(record.get("reference"), 180),
    }

def generation_red_thresholds(flat_records: List[Dict[str, Any]]) -> Dict[str, float]:
    lower_is_worse_metrics = ["rouge_l", "bleu_2", "bertscore_f1"]
    higher_is_worse_metrics = ["ppl", "nll"]

    thresholds = {
        metric_name: mean_std_arith(
            [safe_float(record.get(metric_name)) for record in flat_records]
        )
        for metric_name in lower_is_worse_metrics
    }

    thresholds.update(
        {
            metric_name: mean_std_arith(
                [safe_float(record.get(metric_name)) for record in flat_records],
                subtraction=False
            )
            for metric_name in higher_is_worse_metrics
        }
    )

    return thresholds


def save_generation_rankings_excel(
    records: List[Dict[str, Any]],
    output_path: Path,
    yellow_ratio: float = DEFAULT_YELLOW_RATIO,
) -> None:
    """
    Save generation ranking sheets.

    All generation sheets are sorted ascending because lower generation
    metrics are worse.

    Flag semantics:
    - Red: metric <= mean - std
    - Yellow: user-defined worst percentile
    - Orange: both
ve_generation
    Args:
        records: Unified generation prediction records.
        output_path: Output xlsx path.
        yellow_ratio: User-defined worst percentile ratio.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)

    columns = [
        "rank",
        "record_index",
        "dialog_index",
        "turn_index",
        "bad_consensus_score",
        "bad_consensus_detail",
        "rouge_l",
        "bleu_2",
        "bertscore_f1",
        "nll",
        "ppl",
        "pred_preview",
        "ref_preview",
        "flag",
        "red_threshold",
    ]

    flat_records = [flatten_generation_record(record) for record in records]
    thresholds = generation_red_thresholds(flat_records)

    flat_records = add_bad_consensus_scores(
        records=flat_records,
        thresholds=thresholds,
    )

    metric_to_sheet = {
        "bad_consensus_score": ("by_bad_consensus_desc", False),
        "rouge_l": ("by_rouge_l_asc", True),
        "bleu_2": ("by_bleu_2_asc", True),
        "bertscore_f1": ("by_bertscore_f1_asc", True),
        "nll": ("by_nll_desc", False),
        "ppl": ("by_ppl_desc", False),
    }

    for metric_name, (sheet_name, lower_is_worse) in metric_to_sheet.items():
        sorted_records = sorted(
            flat_records,
            key=lambda r: safe_float(r.get(metric_name)),
            reverse=not lower_is_worse,
        )

        if metric_name == "bad_consensus_score":
            sorted_records = sorted(
                flat_records,
                key=lambda r: (
                    safe_float(r.get("bad_consensus_score")),
                    safe_float(r.get("ppl")),
                    safe_float(r.get("nll")),
                    -safe_float(r.get("rouge_l")),
                    -safe_float(r.get("bleu_2")),
                    -safe_float(r.get("bertscore_f1")),
                ),
                reverse=True,
            )

            flagged_records = [
                {
                    **record,
                    "flag": (
                        "red_bad_consensus"
                        if safe_float(record.get("bad_consensus_score")) > 0
                        else ""
                    ),
                    "red_threshold": None,
                }
                for record in sorted_records
            ]
        else:
            sorted_records = sorted(
                flat_records,
                key=lambda r: safe_float(r.get(metric_name)),
                reverse=not lower_is_worse,
            )

            flagged_records = add_metric_flags(
                records=sorted_records,
                metric_name=metric_name,
                red_threshold=thresholds[metric_name],
                yellow_ratio=yellow_ratio,
                lower_is_worse=lower_is_worse,
            )

        sheet = write_sheet(workbook, sheet_name, add_rank(flagged_records), columns)
        mark_flag_cells(sheet)
        format_numeric_columns(
            sheet,
            [
                "bad_consensus_score",
                "rouge_l",
                "bleu_2",
                "bertscore_f1",
                "nll",
                "ppl",
                "red_threshold",
            ],
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)